"""
backtest.py — Local backtesting engine for the trading strategy.

Usage
─────
  # Test BTC over last 60 days
  python3 backtest.py --coin BTC --days 60

  # Test all coins, last 30 days, save results
  python3 backtest.py --days 30

  # Custom date range
  python3 backtest.py --coin ETH --start 2024-10-01 --end 2024-12-31

  # More capital, different sizing
  python3 backtest.py --balance 5000 --trade-size 50

Output
──────
  • Printed summary table in terminal
  • backtest_results.xlsx — full trade log + equity curve per coin

How it works
────────────
  1. Fetches historical 1H candles from Hyperliquid public API
  2. Walks through each candle in chronological order
  3. At each step: computes all indicators on the candles BEFORE that point
     (no lookahead — we only use past data)
  4. Generates a signal (LONG/SHORT/FLAT)
  5. Simulates entry/exit with planned SL/TP and basic market/limit execution
  6. Tracks portfolio equity, drawdown, win rate, and edges vs simple baselines

Limitations
───────────
  • Uses a local key-level proxy instead of historical L2 replay
  • Sentiment is still neutral unless you extend the evaluator with archived news
  • Execution planning is approximated from candles, not full order-book tape
"""

import sys
import argparse
from datetime import datetime, timedelta
from types import SimpleNamespace
from typing import List, Optional
import pandas as pd
import numpy as np

# ── Load agent modules ────────────────────────────────────────
from config import config
from data.market_data import completed_candle_frame, fetch_candles
from indicators.technical import compute_signals
from indicators.advanced  import compute_advanced_signals
from indicators.candlestick_patterns import compute_candlestick_patterns
from indicators.mtf import _combine as combine_mtf_bias
from indicators.mtf import _compute_bias as compute_timeframe_bias
from indicators.regimes import compute_regimes
from strategy.aggressive_strategy import AggressiveStrategy
from logger import get_logger

log = get_logger("backtest")

NEUTRAL_SENTIMENT = {
    "signal_score": 50.0,
    "raw_score":    50,
    "label":        "Neutral",
    "is_extreme":   False,
}


def _ensure_timestamp_index(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    if "timestamp" in work.columns:
        ts = pd.to_datetime(work["timestamp"], errors="coerce", utc=True)
        if ts.notna().any():
            work.index = ts
            return work
    work.index = pd.date_range(end=datetime.utcnow(), periods=len(work), freq="h")
    return work


def _resample_ohlcv(df: pd.DataFrame, rule: str) -> pd.DataFrame:
    work = _ensure_timestamp_index(df)
    agg = (
        work[["open", "high", "low", "close", "volume"]]
        .resample(rule)
        .agg({
            "open": "first",
            "high": "max",
            "low": "min",
            "close": "last",
            "volume": "sum",
        })
        .dropna()
        .reset_index(drop=True)
    )
    return agg


def _mtf_from_window(coin: str, window: pd.DataFrame):
    one_h = completed_candle_frame(window)
    four_h = completed_candle_frame(_resample_ohlcv(window, "4h"))
    twelve_h = completed_candle_frame(_resample_ohlcv(window, "12h"))
    if one_h is None or len(one_h) < 20:
        return None
    b1 = compute_timeframe_bias(one_h, "1h")
    b4 = compute_timeframe_bias(four_h, "4h") if four_h is not None and len(four_h) >= 20 else None
    b12 = compute_timeframe_bias(twelve_h, "12h") if twelve_h is not None and len(twelve_h) >= 20 else None
    return combine_mtf_bias(coin, b1, b4, b12)


def _local_level_proxy(window: pd.DataFrame, price: float):
    if window is None or window.empty:
        return None
    recent = window.tail(min(len(window), 48))
    if len(recent) < 12:
        return None

    resistance = float(recent["high"].tail(24).max())
    support = float(recent["low"].tail(24).min())
    resistance_distance = abs(resistance - price) / max(price, 1e-9) * 100 if resistance > 0 else 0.0
    support_distance = abs(price - support) / max(price, 1e-9) * 100 if support > 0 else 0.0
    breakout_state = "NONE"
    if resistance > 0 and price > resistance * 1.001:
        breakout_state = "CONFIRMED_BULLISH_BREAKOUT"
    elif support > 0 and price < support * 0.999:
        breakout_state = "CONFIRMED_BEARISH_BREAKDOWN"

    level_interaction = "BETWEEN_LEVELS"
    range_compression = support_distance <= 0.55 and resistance_distance <= 0.55
    if range_compression and breakout_state == "NONE":
        level_interaction = "RANGE_COMPRESSION"
    elif support_distance <= 0.45:
        level_interaction = "AT_SUPPORT"
    elif resistance_distance <= 0.45:
        level_interaction = "AT_RESISTANCE"
    elif support_distance <= 1.25:
        level_interaction = "ABOVE_SUPPORT"
    elif resistance_distance <= 1.25:
        level_interaction = "BELOW_RESISTANCE"
    elif breakout_state == "CONFIRMED_BULLISH_BREAKOUT":
        level_interaction = "ABOVE_BREAKOUT"
    elif breakout_state == "CONFIRMED_BEARISH_BREAKDOWN":
        level_interaction = "BELOW_BREAKDOWN"

    score = 50.0
    if level_interaction in {"AT_SUPPORT", "ABOVE_SUPPORT"}:
        score += 8.0
    if level_interaction in {"AT_RESISTANCE", "BELOW_RESISTANCE"}:
        score -= 8.0
    if breakout_state == "CONFIRMED_BULLISH_BREAKOUT":
        score += 10.0
    elif breakout_state == "CONFIRMED_BEARISH_BREAKDOWN":
        score -= 10.0
    if level_interaction == "RANGE_COMPRESSION":
        score = 50.0 + (score - 50.0) * 0.35

    return SimpleNamespace(
        valid=True,
        score=score,
        level_interaction=level_interaction,
        breakout_state=breakout_state,
        block_longs=level_interaction in {"AT_RESISTANCE", "BELOW_RESISTANCE", "RANGE_COMPRESSION"} and breakout_state != "CONFIRMED_BULLISH_BREAKOUT",
        block_shorts=level_interaction in {"AT_SUPPORT", "ABOVE_SUPPORT", "RANGE_COMPRESSION"} and breakout_state != "CONFIRMED_BEARISH_BREAKDOWN",
        favor_longs=breakout_state == "CONFIRMED_BULLISH_BREAKOUT" or level_interaction in {"AT_SUPPORT", "ABOVE_SUPPORT"},
        favor_shorts=breakout_state == "CONFIRMED_BEARISH_BREAKDOWN" or level_interaction in {"AT_RESISTANCE", "BELOW_RESISTANCE"},
        nearest_support=support,
        nearest_resistance=resistance,
        nearest_support_distance_pct=support_distance,
        nearest_resistance_distance_pct=resistance_distance,
        nearest_support_strength=0.8,
        nearest_resistance_strength=0.8,
        support_levels=[],
        resistance_levels=[],
        imbalance_ratio=0.0,
        bid_notional=1_000_000.0,
        ask_notional=1_000_000.0,
        spread_bps=1.0,
        daily_breakout_level=resistance,
        daily_breakdown_level=support,
    )


def _local_market_map_proxy(window: pd.DataFrame, price: float):
    if window is None or window.empty:
        return None
    daily = _resample_ohlcv(window, "1d")
    daily = completed_candle_frame(daily)
    if daily is None or daily.empty:
        return None

    recent = daily.tail(min(len(daily), 20))
    nearest_support = float(recent["low"].min())
    nearest_resistance = float(recent["high"].max())
    bias = "NEUTRAL"
    score_adjustment = 0.0
    favor_longs = False
    favor_shorts = False
    block_longs = False
    block_shorts = False

    if len(recent) >= 5:
        ema_fast = recent["close"].ewm(span=5, adjust=False).mean().iloc[-1]
        ema_slow = recent["close"].ewm(span=10, adjust=False).mean().iloc[-1]
        if ema_fast > ema_slow and price >= ema_fast:
            bias = "BULLISH"
            score_adjustment = 3.0
            favor_longs = True
        elif ema_fast < ema_slow and price <= ema_fast:
            bias = "BEARISH"
            score_adjustment = -3.0
            favor_shorts = True

    support_distance = abs(price - nearest_support) / max(price, 1e-9) * 100 if nearest_support > 0 else 0.0
    resistance_distance = abs(nearest_resistance - price) / max(price, 1e-9) * 100 if nearest_resistance > 0 else 0.0
    if support_distance <= 0.55:
        favor_longs = True
        block_shorts = True
    if resistance_distance <= 0.55:
        favor_shorts = True
        block_longs = True

    summary = f"{bias} bias with support {nearest_support:.2f} and resistance {nearest_resistance:.2f}"
    return SimpleNamespace(
        valid=True,
        bias=bias,
        score_adjustment=score_adjustment,
        favor_longs=favor_longs,
        favor_shorts=favor_shorts,
        block_longs=block_longs,
        block_shorts=block_shorts,
        nearest_support=nearest_support,
        nearest_resistance=nearest_resistance,
        summary=summary,
        notes="local_proxy",
        daily_close=float(daily["close"].iloc[-1]),
    )


# ─────────────────────────────────────────────────────────────
# Backtest engine
# ─────────────────────────────────────────────────────────────

class Backtester:
    def __init__(
        self,
        coin: str,
        df: pd.DataFrame,
        starting_balance: float = 10_000.0,
        trade_size_usd: float = 20.0,
        stop_loss_pct: float = 0.10,
        take_profit_pct: float = 0.50,
        leverage: int = 3,
    ):
        self.coin = coin
        self.df = df.reset_index(drop=True)
        self.balance = starting_balance
        self.starting_bal = starting_balance
        self.trade_size = trade_size_usd
        self.sl_pct = stop_loss_pct
        self.tp_pct = take_profit_pct
        self.leverage = leverage
        self.strategy = AggressiveStrategy(config.trading, config.indicators)
        self.trades: List[dict] = []
        self.equity_curve: List[float] = []
        self.flat_streak = 0
        self.signal_streak = {"action": None, "count": 0}
        self.pending_entry: Optional[dict] = None

    def _fill_price(self, direction: str, reference_price: float, *, for_exit: bool = False) -> float:
        slip_bps = float(getattr(config.trading, "backtest_market_slippage_bps", 4.0) or 4.0) / 10_000.0
        if direction == "LONG":
            return reference_price * (1 + slip_bps)  # buys fill slightly worse
        return reference_price * (1 - slip_bps)      # sells fill slightly worse

    def _close_position(self, position: dict, exit_price: float, exit_reason: str, closed_at) -> None:
        pnl = self._calc_pnl(position, exit_price)
        self.balance += position["size_usd"] + pnl
        self.trades.append({
            "coin": self.coin,
            "direction": position["direction"],
            "entry_price": position["entry_price"],
            "exit_price": exit_price,
            "size_usd": position["size_usd"],
            "pnl_usd": round(pnl, 2),
            "pnl_pct": round(pnl / position["size_usd"] * 100, 2),
            "exit_reason": exit_reason,
            "opened_at": position["opened_at"],
            "closed_at": str(closed_at),
            "balance_after": round(self.balance, 2),
        })

    def _update_signal_streak(self, action: str) -> bool:
        if action not in {"LONG", "SHORT"}:
            self.signal_streak = {"action": None, "count": 0}
            return False
        if self.signal_streak["action"] == action:
            self.signal_streak["count"] += 1
        else:
            self.signal_streak = {"action": action, "count": 1}
        return self.signal_streak["count"] >= int(getattr(config.trading, "signal_streak_required", 2) or 2)

    def _queue_planned_entry(self, signal, size_usd: float, opened_index: int, opened_at, limit_price: float) -> None:
        trade_plan = dict(getattr(signal, "trade_plan", {}) or {})
        risk_per_unit = float(trade_plan.get("risk_per_unit", 0.0) or 0.0)
        reward_per_unit = float(trade_plan.get("reward_per_unit", 0.0) or 0.0)
        if str(signal.action).upper() == "LONG" and risk_per_unit > 0 and reward_per_unit > 0:
            sl = limit_price - risk_per_unit
            tp = limit_price + reward_per_unit
        elif str(signal.action).upper() == "SHORT" and risk_per_unit > 0 and reward_per_unit > 0:
            sl = limit_price + risk_per_unit
            tp = limit_price - reward_per_unit
        else:
            sl = float(signal.stop_loss_price or 0.0)
            tp = float(signal.take_profit_price or 0.0)
        self.pending_entry = {
            "direction": signal.action,
            "entry_price": limit_price,
            "sl": sl,
            "original_sl": sl,
            "tp": tp,
            "size_usd": size_usd,
            "opened_at": str(opened_at),
            "opened_index": opened_index,
            "score": signal.score,
            "expectancy": dict(getattr(signal, "expectancy", {}) or {}),
            "thesis": dict(getattr(signal, "thesis", {}) or {}),
            "trade_plan": trade_plan,
            "execution_plan": dict(getattr(signal, "execution_plan", {}) or {}),
            "leverage": self.leverage,
            "loss_realization_guard_active": False,
            "wait_cycles": 0,
        }

    def _check_pending_entry_fill(self, current_row, index: int):
        if not self.pending_entry:
            return None

        pending = self.pending_entry
        pending["wait_cycles"] += 1
        hi = float(current_row["high"])
        lo = float(current_row["low"])
        price = float(pending["entry_price"])
        direction = str(pending["direction"])
        fillable = lo <= price <= hi
        max_wait = int(((pending.get("execution_plan") or {}).get("max_wait_cycles")) or 6)

        if fillable:
            self.balance -= pending["size_usd"]
            filled = dict(pending)
            filled["entry_price"] = price
            filled["opened_index"] = index
            self.pending_entry = None
            self.signal_streak = {"action": None, "count": 0}
            return filled
        if pending["wait_cycles"] >= max_wait:
            self.pending_entry = None
        return None

    def _conviction_decay_score(self, position: dict, signal) -> float:
        expectancy = dict(getattr(signal, "expectancy", {}) or {})
        expectancy_score = float(expectancy.get("score", signal.score) or signal.score)
        uncertainty = float(expectancy.get("uncertainty", 0.50) or 0.50)
        expected_r = float(expectancy.get("expected_r", 0.0) or 0.0)
        score = self.flat_streak * float(getattr(config.trading, "conviction_decay_flat_cycle_weight", 7.0) or 7.0)
        if expectancy_score < float(getattr(config.trading, "expectancy_min_score", 56.0) or 56.0):
            score += (float(getattr(config.trading, "expectancy_min_score", 56.0)) - expectancy_score) * 0.9
        if uncertainty >= float(getattr(config.trading, "expectancy_max_uncertainty", 0.42) or 0.42):
            score += (uncertainty - float(getattr(config.trading, "expectancy_max_uncertainty", 0.42))) * 40.0
        if expected_r < float(getattr(config.trading, "expectancy_min_expected_r", 0.18) or 0.18):
            score += float(getattr(config.trading, "conviction_decay_expectancy_weight", 16.0) or 16.0)
        breakout_state = str(getattr(signal, "thesis", {}).get("summary", "") or "")
        if "break" in breakout_state.lower() and signal.action == "FLAT":
            score += 6.0
        return max(0.0, min(100.0, score))

    def _loss_realization_hold_profile(self, position: dict, signal, price: float) -> dict:
        """Backtest mirror of the live thesis-aware loss hold policy."""
        profile = {"defer": False, "hard_stop": float(position.get("sl") or 0.0)}
        if not bool(getattr(config.trading, "loss_realization_guard_enabled", True)):
            return profile

        leverage = max(1.0, float(position.get("leverage") or self.leverage or 1.0))
        max_leverage = float(
            getattr(config.trading, "loss_realization_guard_max_leverage", 3.0) or 3.0
        )
        thesis = dict(getattr(signal, "thesis", {}) or position.get("thesis") or {})
        state = str(thesis.get("state", "") or "").upper()
        conflicts = float(thesis.get("conflict_points", 0.0) or 0.0)
        action = str(getattr(signal, "action", "") or "").upper()
        direction = str(position.get("direction") or "").upper()
        permitted = bool(thesis.get("permitted", action in {"", "FLAT", direction}))
        invalid = state in {"NO_TRADE", "INVALID", "INVALIDATED", "BROKEN"}
        direction_aligned = action in {"", "FLAT", direction}
        scalp_trade = bool(
            position.get("scalp_trade")
            or str((position.get("trade_plan") or {}).get("trade_style", "")).lower() == "scalp"
        )
        thesis_intact = bool(thesis and permitted and direction_aligned and not invalid and conflicts < 2.0)
        if scalp_trade or leverage > max_leverage or not thesis_intact:
            return profile

        entry = float(position.get("entry_price") or 0.0)
        review_stop = float(position.get("original_sl") or position.get("sl") or 0.0)
        risk_distance = abs(entry - review_stop) if entry > 0 and review_stop > 0 else entry * self.sl_pct
        if entry <= 0 or risk_distance <= 0:
            return profile
        favorable_move = price - entry if direction == "LONG" else entry - price
        adverse_r = max(0.0, -favorable_move) / max(risk_distance, 1e-9)
        hard_r = float(
            getattr(config.trading, "loss_realization_guard_hard_adverse_r", 1.75) or 1.75
        )
        hard_stop = entry - risk_distance * hard_r if direction == "LONG" else entry + risk_distance * hard_r
        min_buffer = float(
            getattr(config.trading, "loss_realization_guard_min_liquidation_buffer_pct", 0.12) or 0.12
        )
        if leverage > 1.0:
            if direction == "LONG":
                approximate_liquidation = max(0.0, entry * (1.0 - 1.0 / leverage))
                hard_stop = max(hard_stop, approximate_liquidation + entry * min_buffer)
            else:
                approximate_liquidation = entry * (1.0 + 1.0 / leverage)
                hard_stop = min(hard_stop, approximate_liquidation - entry * min_buffer)

        hard_boundary_hit = bool(
            position.get("loss_realization_guard_active")
            and (
                (direction == "LONG" and price <= float(position.get("sl") or hard_stop))
                or (direction == "SHORT" and price >= float(position.get("sl") or hard_stop))
            )
        )
        risk_room = (direction == "LONG" and price > hard_stop) or (direction == "SHORT" and price < hard_stop)
        if hard_boundary_hit or adverse_r >= hard_r or not risk_room or favorable_move >= 0:
            return profile

        return {
            "defer": True,
            "hard_stop": hard_stop,
            "adverse_r": adverse_r,
            "unrealized_pnl_pct": favorable_move / max(entry, 1e-9) * self.leverage * 100.0,
        }

    def _run_baselines(self) -> dict:
        if self.df.empty:
            return {}

        first_open = float(self.df.iloc[0]["open"])
        last_close = float(self.df.iloc[-1]["close"])
        buy_hold_return = ((last_close - first_open) / max(first_open, 1e-9)) * 100.0

        breakout_position = None
        breakout_balance = self.starting_bal
        dip_position = None
        dip_balance = self.starting_bal

        for i in range(25, len(self.df)):
            window = self.df.iloc[:i].copy()
            row = self.df.iloc[i]
            price = float(row["close"])
            high = float(row["high"])
            low = float(row["low"])
            prev_high = float(window["high"].tail(20).max())
            ema20 = float(window["close"].ewm(span=20, adjust=False).mean().iloc[-1])
            ema50 = float(window["close"].ewm(span=50, adjust=False).mean().iloc[-1])

            if breakout_position is None and price > prev_high * 1.001:
                breakout_position = {"entry": price, "size": self.trade_size}
                breakout_balance -= self.trade_size
            elif breakout_position is not None:
                if low <= breakout_position["entry"] * 0.97 or price < ema20:
                    pnl = ((price - breakout_position["entry"]) / breakout_position["entry"]) * breakout_position["size"] * self.leverage
                    breakout_balance += breakout_position["size"] + pnl
                    breakout_position = None

            dip_condition = ema20 > ema50 and low <= ema20 * 1.002
            if dip_position is None and dip_condition:
                dip_position = {"entry": price, "size": self.trade_size}
                dip_balance -= self.trade_size
            elif dip_position is not None:
                if low <= dip_position["entry"] * 0.97 or price < ema50:
                    pnl = ((price - dip_position["entry"]) / dip_position["entry"]) * dip_position["size"] * self.leverage
                    dip_balance += dip_position["size"] + pnl
                    dip_position = None

        if breakout_position is not None:
            pnl = ((last_close - breakout_position["entry"]) / breakout_position["entry"]) * breakout_position["size"] * self.leverage
            breakout_balance += breakout_position["size"] + pnl
        if dip_position is not None:
            pnl = ((last_close - dip_position["entry"]) / dip_position["entry"]) * dip_position["size"] * self.leverage
            dip_balance += dip_position["size"] + pnl

        return {
            "do_nothing_return": 0.0,
            "buy_hold_return": round(buy_hold_return, 2),
            "breakout_return": round((breakout_balance - self.starting_bal) / max(self.starting_bal, 1e-9) * 100.0, 2),
            "dip_buy_return": round((dip_balance - self.starting_bal) / max(self.starting_bal, 1e-9) * 100.0, 2),
        }

    def run(self, min_candles: int = 80) -> dict:
        df = self.df
        n = len(df)
        position = None

        log.info(f"[{self.coin}] Running walk-forward backtest on {n} candles…")

        for i in range(min_candles, n):
            window = df.iloc[:i].copy()
            analysis_window = completed_candle_frame(window)
            if analysis_window is None or len(analysis_window) < 50:
                continue

            current_row = df.iloc[i]
            entry_reference_price = float(current_row.get("open", current_row["close"]))
            mark_price = float(current_row["close"])
            hi = float(current_row["high"])
            lo = float(current_row["low"])
            ts = current_row.get("timestamp", str(i))

            self.equity_curve.append(self._equity(position, mark_price))

            if position is None:
                filled_entry = self._check_pending_entry_fill(current_row, i)
                if filled_entry is not None:
                    position = filled_entry

            try:
                tech = compute_signals(analysis_window, self.coin, config.indicators, config.trading)
                advanced = compute_advanced_signals(analysis_window, self.coin)
                regimes = compute_regimes(analysis_window, self.coin)
                candle_patterns = compute_candlestick_patterns(analysis_window, self.coin)
                mtf = _mtf_from_window(self.coin, window)
                orderbook_signal = _local_level_proxy(analysis_window, float(analysis_window["close"].iloc[-1]))
                market_map_signal = _local_market_map_proxy(analysis_window, float(analysis_window["close"].iloc[-1]))
            except Exception:
                continue

            if not tech.valid:
                continue

            tech.closed_price = float(analysis_window["close"].iloc[-1])
            tech.live_price = entry_reference_price
            tech.price = entry_reference_price

            signal = self.strategy.generate_signal(
                tech,
                advanced,
                NEUTRAL_SENTIMENT,
                position["direction"] if position else None,
                regimes,
                news_signal=None,
                candle_patterns=candle_patterns,
                memory_adjustment=0.0,
                instrument_type=config.trading.instrument_types.get(self.coin, "crypto"),
                funding_oi_signal=None,
                orderbook_signal=orderbook_signal,
                market_map_signal=market_map_signal,
                narrative_signal=None,
            )

            if mtf:
                if signal.action == "LONG" and not mtf.allow_long:
                    signal.action = "FLAT"
                    signal.flat_reason = f"MTF blocked LONG ({mtf.reason})"
                elif signal.action == "SHORT" and not mtf.allow_short:
                    signal.action = "FLAT"
                    signal.flat_reason = f"MTF blocked SHORT ({mtf.reason})"

            # ── Check exits first using the current candle range ───────────
            if position:
                exit_reason = None
                exit_price = mark_price
                if position["direction"] == "LONG":
                    if lo <= position["sl"]:
                        loss_hold = self._loss_realization_hold_profile(position, signal, float(position["sl"]))
                        if loss_hold.get("defer", False):
                            position["loss_realization_guard_active"] = True
                            position["sl"] = float(loss_hold["hard_stop"])
                            if lo <= position["sl"]:
                                exit_reason = "stop_loss"
                                exit_price = self._fill_price("SHORT", position["sl"], for_exit=True)
                        else:
                            exit_reason = "stop_loss"
                            exit_price = self._fill_price("SHORT", position["sl"], for_exit=True)
                    elif hi >= position["tp"]:
                        exit_reason = "take_profit"
                        exit_price = self._fill_price("SHORT", position["tp"], for_exit=True)
                else:
                    if hi >= position["sl"]:
                        loss_hold = self._loss_realization_hold_profile(position, signal, float(position["sl"]))
                        if loss_hold.get("defer", False):
                            position["loss_realization_guard_active"] = True
                            position["sl"] = float(loss_hold["hard_stop"])
                            if hi >= position["sl"]:
                                exit_reason = "stop_loss"
                                exit_price = self._fill_price("LONG", position["sl"], for_exit=True)
                        else:
                            exit_reason = "stop_loss"
                            exit_price = self._fill_price("LONG", position["sl"], for_exit=True)
                    elif lo <= position["tp"]:
                        exit_reason = "take_profit"
                        exit_price = self._fill_price("LONG", position["tp"], for_exit=True)

                if exit_reason:
                    self._close_position(position, exit_price, exit_reason, ts)
                    position = None
                    self.flat_streak = 0
                    continue

            if position:
                hold_minutes = max(60.0, (i - position["opened_index"]) * 60.0)
                planned_tp = abs(position["tp"] - position["entry_price"])
                favorable_move = (
                    mark_price - position["entry_price"]
                    if position["direction"] == "LONG"
                    else position["entry_price"] - mark_price
                )
                tp_progress = favorable_move / max(planned_tp, 1e-9) if planned_tp > 0 else 0.0

                if signal.action == "FLAT":
                    self.flat_streak += 1
                else:
                    self.flat_streak = 0

                decay_score = self._conviction_decay_score(position, signal)
                loss_hold = self._loss_realization_hold_profile(position, signal, mark_price)
                if (
                    decay_score >= float(getattr(config.trading, "conviction_decay_exit_threshold", 58.0) or 58.0)
                    and not loss_hold.get("defer", False)
                ):
                    exit_price = self._fill_price(
                        "SHORT" if position["direction"] == "LONG" else "LONG",
                        entry_reference_price,
                        for_exit=True,
                    )
                    self._close_position(position, exit_price, "conviction_lost", ts)
                    position = None
                    self.flat_streak = 0
                    continue

                if (
                    hold_minutes >= config.trading.time_stop_minutes
                    and tp_progress < config.trading.time_stop_min_tp_progress
                    and signal.action == "FLAT"
                    and not loss_hold.get("defer", False)
                ):
                    exit_price = self._fill_price(
                        "SHORT" if position["direction"] == "LONG" else "LONG",
                        entry_reference_price,
                        for_exit=True,
                    )
                    self._close_position(position, exit_price, "time_stop", ts)
                    position = None
                    self.flat_streak = 0
                    continue

                if (
                    signal.action in {"LONG", "SHORT"}
                    and signal.action != position["direction"]
                    and hold_minutes >= config.trading.min_hold_minutes
                ):
                    exit_price = self._fill_price(
                        "SHORT" if position["direction"] == "LONG" else "LONG",
                        entry_reference_price,
                        for_exit=True,
                    )
                    self._close_position(position, exit_price, "signal_reversal", ts)
                    position = None
                    self.flat_streak = 0
                    continue

            if position is not None or signal.action == "FLAT":
                continue

            if not self._update_signal_streak(signal.action):
                continue

            size_usd = min(self.trade_size, self.balance * 0.20)
            if size_usd < 1:
                continue

            execution_plan = dict(getattr(signal, "execution_plan", {}) or {})
            if str(execution_plan.get("mode", "market") or "market").lower() in {"limit", "maker_limit"}:
                limit_price = float(execution_plan.get("limit_price", 0.0) or 0.0)
                if limit_price > 0:
                    self._queue_planned_entry(signal, size_usd, i, ts, limit_price)
                    continue

            fill_price = self._fill_price(signal.action, entry_reference_price, for_exit=False)
            sl = float(signal.stop_loss_price or 0.0)
            tp = float(signal.take_profit_price or 0.0)
            if signal.action == "LONG":
                if sl <= 0:
                    sl = fill_price * (1 - self.sl_pct)
                if tp <= 0:
                    tp = fill_price * (1 + self.tp_pct)
            else:
                if sl <= 0:
                    sl = fill_price * (1 + self.sl_pct)
                if tp <= 0:
                    tp = fill_price * (1 - self.tp_pct)

            self.balance -= size_usd
            position = {
                "direction": signal.action,
                "entry_price": fill_price,
                "sl": sl,
                "original_sl": sl,
                "tp": tp,
                "size_usd": size_usd,
                "opened_at": str(ts),
                "opened_index": i,
                "score": signal.score,
                "expectancy": dict(getattr(signal, "expectancy", {}) or {}),
                "thesis": dict(getattr(signal, "thesis", {}) or {}),
                "trade_plan": dict(getattr(signal, "trade_plan", {}) or {}),
                "leverage": self.leverage,
                "loss_realization_guard_active": False,
            }

        if position:
            last_price = self._fill_price(
                "SHORT" if position["direction"] == "LONG" else "LONG",
                float(df.iloc[-1]["close"]),
                for_exit=True,
            )
            self._close_position(position, last_price, "end_of_data", "end")

        return self._summary()

    def _calc_pnl(self, pos: dict, exit_price: float) -> float:
        if pos["direction"] == "LONG":
            pnl_pct = (exit_price - pos["entry_price"]) / pos["entry_price"]
        else:
            pnl_pct = (pos["entry_price"] - exit_price) / pos["entry_price"]
        return pnl_pct * pos["size_usd"] * self.leverage

    def _equity(self, pos: Optional[dict], price: float) -> float:
        if pos:
            return self.balance + pos["size_usd"] + self._calc_pnl(pos, price)
        return self.balance

    def _summary(self) -> dict:
        t = self.trades
        baselines = self._run_baselines()
        if not t:
            return {"coin": self.coin, "trades": 0, "baselines": baselines}

        pnls   = [tr["pnl_usd"] for tr in t]
        wins   = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p <= 0]

        eq     = self.equity_curve
        peak   = self.starting_bal
        max_dd = 0.0
        for e in eq:
            if e > peak:
                peak = e
            dd = (peak - e) / peak * 100
            if dd > max_dd:
                max_dd = dd

        total_return = (self.balance - self.starting_bal) / self.starting_bal * 100
        best_baseline_return = max(baselines.values()) if baselines else 0.0

        return {
            "coin":          self.coin,
            "start_balance": self.starting_bal,
            "end_balance":   round(self.balance, 2),
            "total_return":  round(total_return, 2),
            "trades":        len(t),
            "wins":          len(wins),
            "losses":        len(losses),
            "win_rate":      round(len(wins) / len(t) * 100, 1) if t else 0,
            "total_pnl":     round(sum(pnls), 2),
            "avg_win":       round(sum(wins)   / len(wins)   if wins   else 0, 2),
            "avg_loss":      round(sum(losses) / len(losses) if losses else 0, 2),
            "best_trade":    round(max(pnls), 2),
            "worst_trade":   round(min(pnls), 2),
            "max_drawdown":  round(max_dd, 2),
            "profit_factor": round(sum(wins) / abs(sum(losses)) if losses and sum(losses) != 0 else 0, 2),
            "trade_log":     t,
            "baselines":     baselines,
            "edge_vs_buy_hold": round(total_return - baselines.get("buy_hold_return", 0.0), 2),
            "edge_vs_best_baseline": round(total_return - best_baseline_return, 2),
        }


# ─────────────────────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────────────────────

def save_excel(results: List[dict], path: str = "backtest_results.xlsx"):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    headers = ["Coin", "Trades", "Win Rate %", "Total PnL $", "Return %",
               "Avg Win $", "Avg Loss $", "Best $", "Worst $",
               "Max DD %", "Profit Factor", "End Balance $",
               "BuyHold %", "Breakout %", "DipBuy %", "EdgeVsBest %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")

    green_fill = PatternFill("solid", fgColor="003300")
    red_fill   = PatternFill("solid", fgColor="330000")

    for r in results:
        row = [
            r["coin"], r["trades"],
            r["win_rate"], r["total_pnl"], r["total_return"],
            r["avg_win"], r["avg_loss"], r["best_trade"], r["worst_trade"],
            r["max_drawdown"], r["profit_factor"], r["end_balance"],
            r.get("baselines", {}).get("buy_hold_return", 0.0),
            r.get("baselines", {}).get("breakout_return", 0.0),
            r.get("baselines", {}).get("dip_buy_return", 0.0),
            r.get("edge_vs_best_baseline", 0.0),
        ]
        ws.append(row)
        fill = green_fill if r["total_pnl"] > 0 else red_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 14

    # ── Sheet 2+: Per-coin trade log ──────────────────────────
    for r in results:
        if not r.get("trade_log"):
            continue
        ws2 = wb.create_sheet(title=f"{r['coin']} Trades")
        cols = ["direction", "entry_price", "exit_price", "size_usd",
                "pnl_usd", "pnl_pct", "exit_reason", "opened_at",
                "closed_at", "balance_after"]
        ws2.append(cols)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a1a2e")

        for t in r["trade_log"]:
            ws2.append([t.get(c, "") for c in cols])
            fill = green_fill if t["pnl_usd"] > 0 else red_fill
            for cell in ws2[ws2.max_row]:
                cell.fill = fill

        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 14

    wb.save(path)
    log.info(f"Backtest results saved: {path}")


# ─────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Backtest the trading strategy")
    parser.add_argument("--coin",       default=None,    help="Coin to test (default: all)")
    parser.add_argument("--days",       type=int, default=30, help="Days of history (default: 30)")
    parser.add_argument("--balance",    type=float, default=10_000, help="Starting balance USD")
    parser.add_argument("--trade-size", type=float, default=20,     help="Trade size USD")
    parser.add_argument("--no-excel",   action="store_true",        help="Skip Excel report")
    args = parser.parse_args()

    coins   = [args.coin.upper()] if args.coin else config.trading.coins
    lookback= args.days * 24   # hours → candle count (1H candles)

    all_results = []
    print(f"\n  {'─'*60}")
    print(f"  BACKTEST  |  {args.days} days  |  ${args.balance:,.0f} start  |  ${args.trade_size}/trade")
    print(f"  Coins: {coins}")
    print(f"  {'─'*60}\n")

    for coin in coins:
        print(f"  Fetching {coin} candles ({lookback} periods)…")
        df = fetch_candles(coin=coin, interval="1h", lookback=lookback)
        if df is None or len(df) < 60:
            print(f"  ⚠️  Not enough {coin} data — skipping\n")
            continue

        bt  = Backtester(
            coin              = coin,
            df                = df,
            starting_balance  = args.balance,
            trade_size_usd    = args.trade_size,
            stop_loss_pct     = config.trading.stop_loss_pct,
            take_profit_pct   = config.trading.take_profit_pct,
            leverage          = config.trading.leverage,
        )
        res = bt.run()
        all_results.append(res)

        tag = "✅" if res.get("total_pnl", 0) > 0 else "❌"
        print(f"  {tag} {coin:5s}  Trades={res['trades']:3d}  "
              f"Win={res['win_rate']:5.1f}%  "
              f"PnL=${res['total_pnl']:+,.2f}  "
              f"Return={res['total_return']:+.1f}%  "
              f"MaxDD={res['max_drawdown']:.1f}%  "
              f"EdgeVsBest={res.get('edge_vs_best_baseline', 0.0):+.1f}%")

    print(f"\n  {'─'*60}")
    if all_results:
        total_pnl = sum(r.get("total_pnl", 0) for r in all_results)
        print(f"  Combined PnL: ${total_pnl:+,.2f}")
        if not args.no_excel:
            save_excel(all_results, "backtest_results.xlsx")
            print(f"  Report saved: backtest_results.xlsx")
    print(f"  {'─'*60}\n")


if __name__ == "__main__":
    main()
