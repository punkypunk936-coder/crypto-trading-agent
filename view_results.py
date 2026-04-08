"""
view_results.py — Generate a formatted Excel trade report.

Run this any time to get a fresh snapshot:
    python3 view_results.py

Opens a trades_report.xlsx with two sheets:
  1. Trade Log   — every trade, colour-coded green (win) / red (loss)
  2. Summary     — win rate, PnL, streaks, per-coin breakdown
"""

import csv
import os
import sys
import subprocess
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    print("Installing openpyxl...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

LOG_FILE    = "trades_log.csv"
OUTPUT_FILE = "trades_report.xlsx"

# ── Colour palette ────────────────────────────────────────
C_HEADER_BG  = "1A1A2E"   # dark navy
C_HEADER_FG  = "FFFFFF"   # white
C_WIN_BG     = "D6F5E3"   # soft green
C_LOSS_BG    = "FCE4E4"   # soft red
C_WIN_FG     = "1A6B3C"   # dark green text
C_LOSS_FG    = "9B1C1C"   # dark red text
C_ALT_BG     = "F7F9FC"   # light grey alternating row
C_ACCENT     = "2563EB"   # blue for summary headers
C_SUMMARY_BG = "EFF6FF"   # very light blue for summary cards


def thin_border():
    s = Side(style="thin", color="D0D7E3")
    return Border(left=s, right=s, top=s, bottom=s)


def header_style(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border    = thin_border()


def data_style(cell, bg=None, fg="000000", bold=False, align="center"):
    cell.font      = Font(name="Arial", size=9, color=fg, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def load_trades():
    if not os.path.exists(LOG_FILE):
        return []
    with open(LOG_FILE, newline="") as f:
        return list(csv.DictReader(f))


# ── Sheet 1: Trade Log ────────────────────────────────────

def build_trade_log(wb, trades):
    ws = wb.active
    ws.title = "Trade Log"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:Q1")
    title = ws["A1"]
    title.value     = f"🤖  CRYPTO TRADING AGENT — TRADE LOG   |   Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    title.font      = Font(bold=True, size=12, name="Arial", color=C_HEADER_FG)
    title.fill      = PatternFill("solid", fgColor=C_ACCENT)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Column headers
    cols = [
        ("ID",         5),  ("Coin",     7),  ("Direction", 10),
        ("Opened",    16),  ("Closed",  16),  ("Mins",       7),
        ("Entry $",   11),  ("Exit $",  11),  ("Size $",    10),
        ("Lev",        5),  ("PnL $",   10),  ("PnL %",      8),
        ("Stop $",    10),  ("TP $",    10),  ("Exit Reason",15),
        ("Score",      7),  ("Result",   8),
    ]
    for c, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c, value=label)
        header_style(cell)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[2].height = 22

    # Data rows
    for r, t in enumerate(trades, 3):
        is_win = t.get("result", "LOSS") == "WIN"
        row_bg = C_WIN_BG if is_win else C_LOSS_BG
        row_fg = C_WIN_FG if is_win else C_LOSS_FG
        alt_bg = C_WIN_BG if is_win else C_LOSS_BG

        values = [
            int(t["trade_id"]),
            t["coin"],
            t["direction"],
            t["opened_at"],
            t["closed_at"],
            float(t["duration_mins"]),
            float(t["entry_price"]),
            float(t["exit_price"]),
            float(t["size_usd"]),
            int(t["leverage"]),
            float(t["pnl_usd"]),
            float(t["pnl_pct"]),
            float(t["stop_loss"]),
            float(t["take_profit"]),
            t["exit_reason"].replace("_", " ").title(),
            float(t["signal_score"]),
            t["result"],
        ]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            is_pnl = c in (11, 12)
            data_style(cell, bg=row_bg, fg=row_fg if is_pnl else "2D3748",
                       bold=is_pnl)
            # Number formats
            if c == 11:   cell.number_format = '$#,##0.00;($#,##0.00);"-"'
            elif c == 12: cell.number_format = '+0.00%;-0.00%;"-"'
            elif c in (7, 8, 13, 14): cell.number_format = '$#,##0.00'
            elif c == 9:  cell.number_format = '$#,##0'
            elif c == 6:  cell.number_format = '0.0'

        ws.row_dimensions[r].height = 18

    # Empty state
    if not trades:
        ws.merge_cells("A3:Q3")
        ws["A3"].value     = "No trades yet — run the agent to generate data."
        ws["A3"].font      = Font(italic=True, color="888888", name="Arial")
        ws["A3"].alignment = Alignment(horizontal="center")


# ── Sheet 2: Summary ──────────────────────────────────────

def build_summary(wb, trades):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18

    def card_title(row, col, text, span=2, sheet=ws):
        sheet.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )
        c = sheet.cell(row=row, column=col, value=text)
        c.font      = Font(bold=True, size=10, name="Arial", color=C_HEADER_FG)
        c.fill      = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sheet.row_dimensions[row].height = 22

    def stat_row(row, col, label, formula_or_value, fmt=None, bold_val=False, sheet=ws):
        lc = sheet.cell(row=row, column=col, value=label)
        lc.font      = Font(name="Arial", size=9, color="374151")
        lc.fill      = PatternFill("solid", fgColor=C_SUMMARY_BG)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border    = thin_border()

        vc = sheet.cell(row=row, column=col + 1, value=formula_or_value)
        vc.font      = Font(name="Arial", size=9, color="111827", bold=bold_val)
        vc.fill      = PatternFill("solid", fgColor="FFFFFF")
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = thin_border()
        if fmt:
            vc.number_format = fmt
        sheet.row_dimensions[row].height = 18
        return vc

    # ── Pull data from Trade Log sheet ────────────────────
    tl = "'Trade Log'"     # sheet reference
    n  = len(trades)

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = "📊  PERFORMANCE SUMMARY"
    t.font      = Font(bold=True, size=13, name="Arial", color=C_HEADER_FG)
    t.fill      = PatternFill("solid", fgColor="1A1A2E")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # ── Left column: Overall Stats ────────────────────────
    card_title(3, 1, "  OVERALL PERFORMANCE")

    data_row = n + 3  # last data row in Trade Log (header on row 2, data from row 3)
    last     = data_row if n > 0 else 3

    stat_row(4,  1, "Total Trades",
             f"=COUNTA('{ws.parent['Trade Log'].title}'!A3:A{last+50})-COUNTBLANK('{ws.parent['Trade Log'].title}'!A3:A{last+50})")
    stat_row(5,  1, "Winning Trades",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!Q3:Q{last+50},\"WIN\")")
    stat_row(6,  1, "Losing Trades",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!Q3:Q{last+50},\"LOSS\")")
    stat_row(7,  1, "Win Rate",
             f"=IFERROR(B5/B4,0)", fmt="0.0%", bold_val=True)
    stat_row(8,  1, "Total PnL ($)",
             f"=IFERROR(SUMIF('{ws.parent['Trade Log'].title}'!Q3:Q{last+50},\"WIN\",'{ws.parent['Trade Log'].title}'!K3:K{last+50})+SUMIF('{ws.parent['Trade Log'].title}'!Q3:Q{last+50},\"LOSS\",'{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00;($#,##0.00);"-"', bold_val=True)
    stat_row(9,  1, "Avg Win ($)",
             f"=IFERROR(AVERAGEIF('{ws.parent['Trade Log'].title}'!Q3:Q{last+50},\"WIN\",'{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00')
    stat_row(10, 1, "Avg Loss ($)",
             f"=IFERROR(AVERAGEIF('{ws.parent['Trade Log'].title}'!Q3:Q{last+50},\"LOSS\",'{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00;($#,##0.00)')
    stat_row(11, 1, "Best Trade ($)",
             f"=IFERROR(MAX('{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00', bold_val=True)
    stat_row(12, 1, "Worst Trade ($)",
             f"=IFERROR(MIN('{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00;($#,##0.00)', bold_val=True)
    stat_row(13, 1, "Avg Duration (mins)",
             f"=IFERROR(AVERAGE('{ws.parent['Trade Log'].title}'!F3:F{last+50}),0)",
             fmt='0.0')

    # ── Right column: Exit Breakdown ─────────────────────
    card_title(3, 4, "  EXIT BREAKDOWN")
    stat_row(4,  4, "Take Profits Hit",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!O3:O{last+50},\"Take Profit\")")
    stat_row(5,  4, "Stop Losses Hit",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!O3:O{last+50},\"Stop Loss\")")
    stat_row(6,  4, "Trailing Stops Hit",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!O3:O{last+50},\"Trailing Stop\")")
    stat_row(7,  4, "Signal Reversals",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!O3:O{last+50},\"Signal Reversal\")")

    # ── Right column lower: Per-coin ─────────────────────
    card_title(10, 4, "  PER-COIN PnL ($)")
    for i, coin in enumerate(["BTC", "ETH", "SOL", "HYPE"], 11):
        stat_row(i, 4, coin,
                 f"=IFERROR(SUMIF('{ws.parent['Trade Log'].title}'!B3:B{last+50},\"{coin}\",'{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
                 fmt='$#,##0.00;($#,##0.00);"-"')

    # ── Direction breakdown ───────────────────────────────
    card_title(16, 1, "  LONG vs SHORT")
    stat_row(17, 1, "Long Trades",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!C3:C{last+50},\"LONG\")")
    stat_row(18, 1, "Short Trades",
             f"=COUNTIF('{ws.parent['Trade Log'].title}'!C3:C{last+50},\"SHORT\")")
    stat_row(19, 1, "Long PnL ($)",
             f"=IFERROR(SUMIF('{ws.parent['Trade Log'].title}'!C3:C{last+50},\"LONG\",'{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00;($#,##0.00);"-"')
    stat_row(20, 1, "Short PnL ($)",
             f"=IFERROR(SUMIF('{ws.parent['Trade Log'].title}'!C3:C{last+50},\"SHORT\",'{ws.parent['Trade Log'].title}'!K3:K{last+50}),0)",
             fmt='$#,##0.00;($#,##0.00);"-"')


# ── Main ──────────────────────────────────────────────────

def generate_report():
    trades = load_trades()

    wb = Workbook()
    build_trade_log(wb, trades)
    build_summary(wb, trades)

    wb.save(OUTPUT_FILE)
    print(f"\n✅  Report saved: {OUTPUT_FILE}")
    print(f"   {len(trades)} trade(s) logged\n")

    # Auto-open on Mac
    try:
        subprocess.run(["open", OUTPUT_FILE], check=False)
    except Exception:
        pass


if __name__ == "__main__":
    generate_report()
